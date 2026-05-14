"""
JurisFlow — Planilha de Custos
2 abas: Custo por Token | Infraestrutura
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"c:\Users\lakil\Downloads\PROJETO API-CONTESTACAO\API-CONTESTACAO\docs\JurisFlow_Custos.xlsx"

AZUL_ESC   = "1E3A5F"
AZUL_MED   = "2E6DA4"
AZUL_FUNDO = "EBF3FB"
VERDE_ESC  = "217346"
VERDE_CLR  = "C6EFCE"
BRANCO     = "FFFFFF"
CINZA      = "F2F2F2"
VERMELHO   = "C00000"

BRL            = 5.10
CUSTO_UNIT_USD = 0.1003

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width
def h(ws, row, height):
    ws.row_dimensions[row].height = height

def hdr(ws, row, col, text, bg=AZUL_ESC, fg=BRANCO, size=9, colspan=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=fg, size=size)
    c.fill = fill(bg)
    c.alignment = align("center")
    c.border = border()
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    return c

def cel(ws, row, col, value=None, bold=False, color="000000", bg=BRANCO,
        ha="left", fmt=None, size=9, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=color, size=size, italic=italic)
    c.fill = fill(bg)
    c.alignment = align(ha, wrap=True)
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

def secao(ws, row, col, text, end_col, bg=AZUL_MED):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=BRANCO, size=10)
    c.fill = fill(bg)
    c.alignment = align("left")
    c.border = border()
    h(ws, row, 20)

def titulo(ws, row, text, ncols, sub=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(bold=True, color=BRANCO, size=14)
    c.fill = fill(AZUL_ESC)
    c.alignment = align("center")
    h(ws, row, 34)
    if sub:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
        s = ws.cell(row=row+1, column=1, value=sub)
        s.font = font(italic=True, color=AZUL_MED, size=9)
        s.fill = fill(AZUL_FUNDO)
        s.alignment = align("center")
        h(ws, row+1, 18)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ABA 1 — Custo por Token
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def aba_tokens(wb):
    ws = wb.create_sheet("Custo por Token")
    ws.sheet_view.showGridLines = False
    NC = 2

    for col, width in [(1,36),(2,24)]:
        w(ws, col, width)

    titulo(ws, 1, "JurisFlow — Custo por Contestação (API Claude Sonnet 4.6)",
           NC, "Câmbio R$ 5,10  |  Modelo: Claude Sonnet 4.6")

    # valores internos para cálculo
    total_in = 19192
    tot_min  = 2070
    tot_typ  = 3480
    tot_max  = 5850
    cached   = 3107
    fresh    = total_in - cached

    def custo(out, cache=False):
        if cache:
            return (fresh * 3 + cached * 0.30 + out * 15) / 1_000_000
        return (total_in * 3 + out * 15) / 1_000_000

    # ── A: Custo por contestação ─────────────────────────────────────────
    r = 4
    secao(ws, r, 1, "A   CUSTO POR CONTESTAÇÃO — cenários", NC)

    r += 1
    hdr(ws, r, 1, "Cenário",           bg=AZUL_MED)
    hdr(ws, r, 2, "Custo BRL (R$ 5,10)", bg=AZUL_MED)
    h(ws, r, 18)

    cenarios = [
        ("Sem cache — mínimo",   custo(tot_min)),
        ("Sem cache — típico",   custo(tot_typ)),
        ("Sem cache — máximo",   custo(tot_max)),
        ("Com cache — mínimo",   custo(tot_min, True)),
        ("Com cache — típico ★", custo(tot_typ, True)),
        ("Com cache — máximo",   custo(tot_max, True)),
    ]

    for i, (cen, usd) in enumerate(cenarios):
        star = "★" in cen
        bg   = VERDE_CLR if star else (BRANCO if i % 2 == 0 else AZUL_FUNDO)
        cel(ws, r+1+i, 1, cen,       bold=star, bg=bg)
        cel(ws, r+1+i, 2, usd * BRL, bold=star, bg=bg, ha="center", fmt='"R$" #,##0.0000')
        h(ws, r+1+i, 17)

    r += len(cenarios) + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    nota = ws.cell(row=r, column=1,
        value="★ Cenário de referência  |  Com cache: system prompt e template são reutilizados entre requisições  |  "
              "Preços Anthropic: Input $3/1M · Output $15/1M · Cache read $0,30/1M")
    nota.font = font(italic=True, color=AZUL_MED, size=8)
    nota.fill = fill(AZUL_FUNDO)
    nota.alignment = align("left", wrap=True)
    nota.border = border()
    h(ws, r, 30)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ABA 2 — Infraestrutura
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def aba_infra(wb):
    ws = wb.create_sheet("Infraestrutura")
    ws.sheet_view.showGridLines = False
    NC = 6

    for col, width in [(1,22),(2,14),(3,14),(4,14),(5,28),(6,28)]:
        w(ws, col, width)

    titulo(ws, 1, "JurisFlow — Custos de Infraestrutura", NC,
           "Todos os serviços que sustentam a plataforma — planos iniciais e limites para upgrade")

    # ── A: Plano inicial ──────────────────────────────────────────────────
    r = 4
    secao(ws, r, 1, "A   PLANO INICIAL — Free / Starter", NC)

    r += 1
    for ci, txt in enumerate(["Serviço","Plano","USD/mês","BRL/mês","Limites","Função no JurisFlow"], 1):
        hdr(ws, r, ci, txt, bg=AZUL_MED)
    h(ws, r, 18)

    servicos = [
        ("Supabase",      "Free",     0,    0,           "500 MB DB · 50k MAU · auth ilimitada",   "Banco de dados PostgreSQL + autenticação"),
        ("n8n",           "Self-host",0,    0,           "Ilimitado (roda dentro do Railway)",      "Orquestração de workflows e chamadas à IA"),
        ("Railway",       "Starter",  5,    5*BRL,       "8 GB RAM · 8 vCPU · $5 crédito incluso", "Hospeda Backend FastAPI + n8n (Docker Compose)"),
        ("Vercel",        "Hobby",    0,    0,           "100 GB banda · deploys ilimitados",       "Hospedagem do frontend React/Vite"),
        ("Anthropic API", "Pay-as-go",None, None,        "Sem mensalidade · sem mínimo",            "API Claude Sonnet 4.6 (custo variável)"),
        ("GitHub",        "Free",     0,    0,           "Repos privados ilimitados",               "Controle de versão + CI/CD automático"),
    ]

    for i, (srv, plano, usd, brl, lim, func) in enumerate(servicos):
        bg = BRANCO if i % 2 == 0 else AZUL_FUNDO
        cel(ws, r+1+i, 1, srv,   bold=True, bg=bg)
        cel(ws, r+1+i, 2, plano, bg=bg, ha="center")
        if usd is None:
            cel(ws, r+1+i, 3, "variável", bg=bg, ha="center", italic=True, color="666666")
            cel(ws, r+1+i, 4, "variável", bg=bg, ha="center", italic=True, color="666666")
        else:
            cel(ws, r+1+i, 3, usd, bg=bg, ha="center", fmt='"$" #,##0.00')
            cel(ws, r+1+i, 4, brl, bg=bg, ha="center", fmt='"R$" #,##0.00')
        cel(ws, r+1+i, 5, lim,  bg=bg, color="444444", italic=True)
        cel(ws, r+1+i, 6, func, bg=bg)
        h(ws, r+1+i, 19)

    r += len(servicos) + 1
    cel(ws, r, 1, "TOTAL FIXO/MÊS", bold=True, color=BRANCO, bg=AZUL_ESC, ha="right")
    cel(ws, r, 2, "", color=BRANCO, bg=AZUL_ESC)
    cel(ws, r, 3, 5.0, bold=True, color=BRANCO, bg=AZUL_ESC, ha="center", fmt='"$" #,##0.00')
    cel(ws, r, 4, 5*BRL, bold=True, color=BRANCO, bg=AZUL_ESC, ha="center", fmt='"R$" #,##0.00')
    cel(ws, r, 5, "Apenas Railway Starter (Supabase, Vercel e GitHub são gratuitos)",
        color=BRANCO, bg=AZUL_ESC)
    cel(ws, r, 6, "", color=BRANCO, bg=AZUL_ESC)
    h(ws, r, 20)

    # ── B: Quando migrar ──────────────────────────────────────────────────
    r += 2
    secao(ws, r, 1, "B   QUANDO MIGRAR PARA PLANOS PAGOS", NC, bg=VERMELHO)

    r += 1
    for ci, txt in enumerate(["Serviço","Próximo plano","USD/mês","BRL/mês","Gatilho para upgrade","Benefício principal"], 1):
        hdr(ws, r, ci, txt, bg=VERMELHO)
    h(ws, r, 18)

    upgrades = [
        ("Supabase",      "Pro",        25,   25*BRL,    "> 500 MB DB ou > 50k usuários/mês",       "8 GB DB, backups diários, suporte e-mail"),
        ("n8n",           "Cloud Starter",20, 20*BRL,   "Precisa de interface sem gerenciar servidor","UI visual gerenciada, SSL, atualizações auto"),
        ("Railway",       "Pro",        20,   20*BRL,    "> 8 GB RAM ou múltiplos serviços pesados", "Escalonamento automático, mais containers"),
        ("Vercel",        "Pro",        20,   20*BRL,    "> 100 GB banda ou equipe com > 1 pessoa",  "Analytics, proteção DDoS, previews ilimitadas"),
        ("Anthropic",     "Committed",  None, None,      "> $5.000/mês em consumo de API",           "Desconto 20–30% negociado por volume"),
    ]

    for i, (srv, plano, usd, brl, gatilho, ben) in enumerate(upgrades):
        bg = BRANCO if i % 2 == 0 else AZUL_FUNDO
        cel(ws, r+1+i, 1, srv,     bold=True, bg=bg)
        cel(ws, r+1+i, 2, plano,   bg=bg, ha="center")
        if usd is None:
            cel(ws, r+1+i, 3, "negociável", bg=bg, ha="center", italic=True, color="666666")
            cel(ws, r+1+i, 4, "negociável", bg=bg, ha="center", italic=True, color="666666")
        else:
            cel(ws, r+1+i, 3, usd, bg=bg, ha="center", fmt='"$" #,##0.00')
            cel(ws, r+1+i, 4, brl, bg=bg, ha="center", fmt='"R$" #,##0.00')
        cel(ws, r+1+i, 5, gatilho, bg=bg, color="444444")
        cel(ws, r+1+i, 6, ben,     bg=bg)
        h(ws, r+1+i, 19)

    # ── C: Custo total por fase ───────────────────────────────────────────
    r += len(upgrades) + 2
    secao(ws, r, 1, "C   CUSTO TOTAL MENSAL POR FASE (Infra + API Claude)", NC)

    r += 1
    for ci, txt in enumerate(["Fase","Contest./mês","Infra BRL","API Claude BRL","Total BRL/mês","Custo por peça BRL"], 1):
        hdr(ws, r, ci, txt, bg=AZUL_MED)
    h(ws, r, 18)

    fases = [
        ("Lançamento / MVP",    10,   5*BRL),
        ("Crescimento inicial", 50,   5*BRL),
        ("Escritório ativo ★", 100,  5*BRL),
        ("Expansão",           500,  (5+25)*BRL),
        ("Escala",            1000,  (5+25+20)*BRL),
    ]

    for i, (fase, vol, infra) in enumerate(fases):
        api   = vol * CUSTO_UNIT_USD * BRL
        total = infra + api
        star  = "★" in fase
        bg    = VERDE_CLR if star else (BRANCO if i % 2 == 0 else AZUL_FUNDO)
        cel(ws, r+1+i, 1, fase,        bold=star, bg=bg)
        cel(ws, r+1+i, 2, vol,         bg=bg, ha="center")
        cel(ws, r+1+i, 3, infra,       bg=bg, ha="center", fmt='"R$" #,##0.00')
        cel(ws, r+1+i, 4, api,         bg=bg, ha="center", fmt='"R$" #,##0.00')
        cel(ws, r+1+i, 5, total,       bold=star, bg=bg, ha="center", fmt='"R$" #,##0.00')
        cel(ws, r+1+i, 6, total/vol,   bg=bg, ha="center", fmt='"R$" #,##0.00')
        h(ws, r+1+i, 17)

    r += len(fases) + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    nota = ws.cell(row=r, column=1,
        value="★ Fase de referência: 100 contestações/mês  |  "
              "Expansão: inclui Supabase Pro ($25)  |  "
              "Escala: inclui Supabase Pro + Railway Pro (+$45)")
    nota.font = font(italic=True, color=AZUL_MED, size=8)
    nota.fill = fill(AZUL_FUNDO)
    nota.alignment = align("left", wrap=True)
    nota.border = border()
    h(ws, r, 22)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    aba_tokens(wb)
    aba_infra(wb)

    wb["Custo por Token"].sheet_properties.tabColor  = AZUL_ESC
    wb["Infraestrutura"].sheet_properties.tabColor   = VERDE_ESC

    wb.save(OUT)
    print(f"Gerado: {OUT}")

if __name__ == "__main__":
    main()
